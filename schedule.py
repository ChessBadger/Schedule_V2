import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, timedelta
import re
import json
import time
import openpyxl

# Load credentials from a JSON file
with open('config.json') as config_file:
    credentials = json.load(config_file)

# Set up the credentials
scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']
creds = Credentials.from_service_account_info(credentials, scopes=scope)
client = gspread.authorize(creds)

sheet_names = ['Week1', 'Week2']

# Open the Google Sheets document by its title
spreadsheet = client.open(sheet_names[0])


# Select the worksheet by title
worksheet = spreadsheet.sheet1

# Get the value of cell A16
sunday_value = worksheet.acell('A1').value

# Get the current year
current_year = datetime.now().year

# Convert the string to a datetime object with the current year
sunday_date = datetime.strptime(
    sunday_value + f", {current_year}", '%a, %b %d, %Y')

# Create an array with the days of the week
days_of_week = [sunday_date.strftime('%a, %b %d')]

# Add the next two weeks
for i in range(1, 14):
    next_day = sunday_date + timedelta(days=i)
    days_of_week.append(next_day.strftime('%a, %b %d'))


# Create a dictionary with the days of the week as keys and empty lists as values
schedule = {day: [] for day in days_of_week}


class Store_Run:
    # Constructor (initializer) method
    def __init__(self, meet_time, start_time, inv_type, store_name, store_address, store_link, note, store_crew, store_supervisor, store_drivers, is_driver, is_supervisor):
        self.meet_time = meet_time
        self.start_time = start_time
        self.inv_type = inv_type
        self.store_name = store_name
        self.store_address = store_address
        self.store_link = store_link
        self.note = note
        self.store_crew = store_crew
        self.store_supervisor = store_supervisor
        self.store_drivers = store_drivers
        self.is_driver = is_driver
        self.is_supervisor = is_supervisor


def process_employee(employee_name, column_number, counter, excel_file):
    # # Convert employee_name to uppercase
    # employee_name = employee_name.upper

    try:

        is_driver = False
        is_supervisor = False

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active

        # Iterate through the rows in the specified column
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=column_number, max_col=column_number):
            for cell in row:
                # Check for occurrences of employee_name in the column
                if cell.value and cell.value.upper() == employee_name:

                    store_link = []
                    store_address = []
                    store_name = []
                    inv_type = []
                    store_crew = []
                    store_drivers = []

                    # Step 2: Move up one cell at a time until a link is found
                    current_cell = cell

                    # Add the any notes next to name
                    if worksheet.cell(current_cell.row, current_cell.column + 1).value:
                        note = worksheet.cell(
                            current_cell.row, current_cell.column + 1).value
                        # Check if the note contains the word "DRIVER"
                        if "DRIVER" in note.upper():
                            is_driver = True
                    else:
                        note = "None"

                    # Remove new line characters
                    if note:
                        note = note.replace("\n", " ")

                    # Check the employee is supervisor
                    number_cell = worksheet.cell(
                        current_cell.row, current_cell.column - 1)
                    if number_cell.value is not None and number_cell.value == "1)":
                        is_supervisor = True
                        store_supervisor = worksheet.cell(
                            current_cell.row, current_cell.column).value
                    else:
                        store_supervisor = "None"

                    while current_cell.row > 1:
                        current_cell = worksheet.cell(
                            current_cell.row - 1, column_number)
                        # Check if the content of the cell is not None and contains the word "OFFICE"
                        if current_cell.value and "MILWAUKEE OFFICE" in current_cell.value.upper():
                            store_link.append(
                                current_cell.value.replace("\n", " "))
                            store_name.append(
                                current_cell.value.replace("\n", " "))
                            store_address.append("None")
                            inv_type.append("None")
                            start_time = "None"
                            meet_time = "None"
                            break
                        # Check if the content of the cell is not None and is a hyperlink
                        elif current_cell.value and re.match(r'^https?://', current_cell.value):
                            store_link.insert(
                                0, current_cell.value.replace("\n", " "))
                            break

                    else:
                        store_supervisor = "None"

                    if store_link and "OFFICE" not in store_link[0].upper():
                        # Step 3: Move up once cell above the link and set that as the store_address
                        store_address.insert(0, worksheet.cell(
                            current_cell.row - 1, column_number).value.replace("\n", " "))

                        # Step 4: Move up one cell from the store_address and set that cell as the store_name
                        store_name.insert(0, worksheet.cell(
                            current_cell.row - 2, column_number).value.replace("\n", " "))

                        # Step 5: Move one cell up from the store_name and set that cell as the inv_type
                        inv_type.insert(0, worksheet.cell(
                            current_cell.row - 3, column_number).value.replace("\n", " "))

                        # Step 6: Move up one cell at a time until a cell is found that starts with a time in the format of HH:MM or H:MM
                        while current_cell.row > 1:
                            current_cell = worksheet.cell(
                                current_cell.row - 1, column_number)
                            cell_value = current_cell.value
                            # Check for HH:MM format
                            if cell_value and re.match(r'\d{1,2}:\d{2}', cell_value.strip()):
                                start_time = cell_value.strip()
                                break
                            elif current_cell.value and re.match(r'^https?://', current_cell.value):
                                # If another link is found before a time is found, add that link to store_link and repeat steps 3-5
                                store_link.insert(
                                    0, cell_value.replace("\n", " "))
                                store_address.insert(0, worksheet.cell(
                                    current_cell.row - 1, column_number).value.replace("\n", " "))
                                store_name.insert(0, worksheet.cell(
                                    current_cell.row - 2, column_number).value.replace("\n", " "))
                                inv_type.insert(0, worksheet.cell(
                                    current_cell.row - 3, column_number).value.replace("\n", " "))

                    # Step 7: Move up one cell from the start_time and set it to the meet_time
                    if start_time:
                        if worksheet.cell(current_cell.row - 1, column_number).value:
                            meet_time = worksheet.cell(
                                current_cell.row - 1, column_number).value
                        else:
                            meet_time = "None"
                    else:
                        meet_time = "None"

                    # Display crew if employee is supervisor
                    if is_supervisor:
                        current_cell = cell
                        supervisor_cell = current_cell
                        crew_counter = 0
                        while current_cell.row < 130:
                            current_cell = worksheet.cell(
                                current_cell.row + 1, column_number)
                            if current_cell.value:
                                crew_counter += 1
                                employee = str(crew_counter) + ".) " + \
                                    current_cell.value
                                store_crew.append(employee)
                            else:
                                break

                    # Get supervisor if employee is not supervisor
                    elif not is_supervisor and "OFFICE" not in store_link[0].upper():
                        current_cell = cell
                        while current_cell.row > 1:
                            current_cell = worksheet.cell(
                                current_cell.row - 1, column_number - 1)
                            if current_cell.value == "1)":
                                current_cell = worksheet.cell(
                                    current_cell.row, column_number)
                                store_supervisor = current_cell.value
                                supervisor_cell = current_cell
                                break

                    # Display drivers if there is a meet
                    if "None" not in meet_time:
                        while supervisor_cell.row < 130:
                            supervisor_cell = worksheet.cell(
                                supervisor_cell.row + 1, column_number)
                            if supervisor_cell.value:
                                driver_cell = worksheet.cell(
                                    supervisor_cell.row, column_number + 1)
                                if driver_cell.value and "DRIVER" in driver_cell.value.upper():
                                    if employee_name not in supervisor_cell.value.upper():
                                        store_drivers.append(
                                            supervisor_cell.value)
                            else:
                                break

                    # Display crew if employee is driver
                    if "None" not in meet_time:
                        if is_driver and not is_supervisor:
                            crew_counter = 0
                            while current_cell.row < 130:
                                if current_cell.value:
                                    at_store_cell = worksheet.cell(
                                        current_cell.row, column_number + 1).value
                                    if at_store_cell:
                                        if "@ STORE" not in at_store_cell.upper() and "DRIVER" not in at_store_cell.upper():
                                            if employee_name not in current_cell.value:
                                                crew_counter += 1
                                                employee = str(crew_counter) + ".) " + \
                                                    current_cell.value
                                                store_crew.append(employee)
                                    else:
                                        if employee_name not in current_cell.value:
                                            crew_counter += 1
                                            employee = str(crew_counter) + ".) " + \
                                                current_cell.value
                                            store_crew.append(employee)
                                else:
                                    break
                                current_cell = worksheet.cell(
                                    current_cell.row + 1, column_number)

                    # Create an instance of the Store_Run class
                    store_run_instance = Store_Run(
                        meet_time, start_time, inv_type, store_name, store_address, store_link, note, store_crew, store_supervisor, store_drivers, is_driver, is_supervisor)

                    # Append the data to the respective day's entry in the schedule_data dictionary
                    schedule[days_of_week[counter]
                             ].append(store_run_instance.__dict__)
                    print(days_of_week[counter], "Done")

    except gspread.exceptions.APIError as api_error:
        if api_error.response.status_code == 429:  # Rate limit exceeded
            print(f"Rate limit exceeded. Waiting and retrying...")
            time.sleep(60)  # Sleep for 1 minute (adjust as needed)
            # Retry the operation
            process_employee(employee_name, column_number, counter)
        else:
            # Handle other API errors
            print(f"API Error: {api_error}")
    except Exception as e:
        # Handle other exceptions
        print(f"An unexpected error occurred: {e}")


# Update the schedule data in a JSON file after each successful API call
def update_schedule_json(schedule):
    with open('schedule_data.json', 'w') as json_file:
        json.dump(schedule, json_file)


employee_name = input("Enter employee name: ").upper()

# Set which columns to check for employee names
columns_to_process = [2, 6, 10, 14, 18, 22, 26]
counter = 0


for sheet_name in sheet_names:
    # Open the Google Sheets document by its title
    spreadsheet = client.open(sheet_name)
    # Select the worksheet by title
    worksheet = spreadsheet.sheet1
    # Iterate through each column

    # Extract data from Google Sheets
    data = worksheet.get_all_values()

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Insert data into the Excel sheet maintaining formatting
    for row_index, row_data in enumerate(data, start=1):
        for col_index, cell_data in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index).value = cell_data

    workbook.save(sheet_name + '.xlsx')

    for column in columns_to_process:
        try:
            process_employee(employee_name, column, counter,
                             sheet_name + '.xlsx')
            counter += 1
            # Update JSON after each successful API call
            update_schedule_json(schedule)
        except gspread.exceptions.CellNotFound:
            print(f"{employee_name} not found in column {column}")
        except Exception as e:
            print(f"An error occurred: {e}")


# If needed, update JSON after all iterations
update_schedule_json(schedule)
